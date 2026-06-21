"""
backend/export_report.py
Full-dashboard reporting: PDF, Excel (multi-sheet), CSV, HTML.
Every format contains ALL dashboard sections:
  Upload Info → Preview (sample) → Descriptive Stats (Num + Cat)
  → Auto Insights → Time Series summary
"""

import os, json, html, io
from datetime import datetime
from matplotlib import colors
from matplotlib import colors
from matplotlib.pyplot import xlabel, ylabel
import pandas as pd
import numpy as np

# ─── palette ──────────────────────────────────────────────────────────────────
TEAL  = "#0F6E8C"   # aksen biru-teal gelap untuk judul & garis
INK   = "#1A1A1A"   # teks utama (hampir hitam)
INK2  = "#F4F8FA"   # baris tabel genap (abu-abu sangat muda)
INK3  = "#E7EFF2"   # baris tabel ganjil
MINT  = "#1A1A1A"   # teks isi tabel -> gelap (dulunya nyaris putih)
TEA   = "#1F2937"   # sub-judul / caption -> abu gelap
GREEN = "#16A34A"
AMBER = "#D97706"
RED   = "#DC2626"
TEA   = "#CAFFDE"
GREEN = "#22c55e"
AMBER = "#f59e0b"
RED   = "#ef4444"

# Simbol teks netral (bukan emoticon) untuk merepresentasikan tipe insight
# di output statis seperti PDF/Excel/CSV, yang tidak bisa merender SVG.
INSIGHT_TYPE_SYMBOL = {
    "info"   : "i",
    "success": "OK",
    "warning": "!",
    "danger" : "!!",
}


def insight_symbol(ins: dict) -> str:
    """Simbol teks pendek berdasarkan tipe insight, pengganti emoji/icon-id."""
    return INSIGHT_TYPE_SYMBOL.get(ins.get("type", "info"), "i")


# Mapping icon-id (dari InsightGenerator) ke SVG inline, dipakai khusus
# pada HTML report (yang sanggup merender SVG, tidak seperti PDF/Excel/CSV).
INSIGHT_SVG = {
    "trend-up": '<svg viewBox="0 0 24 24" width="18" height="18" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><polyline points="23 6 13.5 15.5 8.5 10.5 1 18"/><polyline points="17 6 23 6 23 12"/></svg>',
    "alert-triangle": '<svg viewBox="0 0 24 24" width="18" height="18" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M10.29 3.86L1.82 18a2 2 0 0 0 1.71 3h16.94a2 2 0 0 0 1.71-3L13.71 3.86a2 2 0 0 0-3.42 0z"/><line x1="12" y1="9" x2="12" y2="13"/><line x1="12" y1="17" x2="12.01" y2="17"/></svg>',
    "check-circle": '<svg viewBox="0 0 24 24" width="18" height="18" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M22 11.08V12a10 10 0 1 1-5.93-9.14"/><polyline points="22 4 12 14.01 9 11.01"/></svg>',
    "alert-circle": '<svg viewBox="0 0 24 24" width="18" height="18" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><circle cx="12" cy="12" r="10"/><line x1="12" y1="8" x2="12" y2="12"/><line x1="12" y1="16" x2="12.01" y2="16"/></svg>',
    "bar-chart": '<svg viewBox="0 0 24 24" width="18" height="18" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><line x1="12" y1="20" x2="12" y2="10"/><line x1="18" y1="20" x2="18" y2="4"/><line x1="6" y1="20" x2="6" y2="16"/></svg>',
    "link": '<svg viewBox="0 0 24 24" width="18" height="18" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M10 13a5 5 0 0 0 7.54.54l3-3a5 5 0 0 0-7.07-7.07l-1.72 1.71"/><path d="M14 11a5 5 0 0 0-7.54-.54l-3 3a5 5 0 0 0 7.07 7.07l1.71-1.71"/></svg>',
    "bell": '<svg viewBox="0 0 24 24" width="18" height="18" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M18 8A6 6 0 0 0 6 8c0 7-3 9-3 9h18s-3-2-3-9"/><path d="M13.73 21a2 2 0 0 1-3.46 0"/></svg>',
    "clock": '<svg viewBox="0 0 24 24" width="18" height="18" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><circle cx="12" cy="12" r="10"/><polyline points="12 6 12 12 16 14"/></svg>',
    "star": '<svg viewBox="0 0 24 24" width="18" height="18" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><polygon points="12 2 15.09 8.26 22 9.27 17 14.14 18.18 21.02 12 17.77 5.82 21.02 7 14.14 2 9.27 8.91 8.26 12 2"/></svg>',
    "filter": '<svg viewBox="0 0 24 24" width="18" height="18" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><polygon points="22 3 2 3 10 12.46 10 19 14 21 14 12.46 22 3"/></svg>',
}
INSIGHT_SVG_DEFAULT = '<svg viewBox="0 0 24 24" width="18" height="18" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><circle cx="12" cy="12" r="10"/><line x1="12" y1="16" x2="12" y2="12"/><line x1="12" y1="8" x2="12.01" y2="8"/></svg>'


def insight_svg(ins: dict) -> str:
    """SVG inline untuk insight card di HTML report (icon, bukan emoticon)."""
    return INSIGHT_SVG.get(ins.get("icon", ""), INSIGHT_SVG_DEFAULT)


class ExportReport:
    def __init__(self, df: pd.DataFrame, session: dict):
        self.df      = df
        self.session = session
        self.ts_str  = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.now     = datetime.now().strftime("%d %b %Y %H:%M")

    # ── helpers ───────────────────────────────────────────────────────────────
    def _num_stats(self):
        return self.session.get("stats", {}).get("numerical", [])

    def _cat_stats(self):
        return self.session.get("stats", {}).get("categorical", [])

    def _insights(self):
        return self.session.get("insights", [])

    def _ts(self):
        return self.session.get("ts", {})

    def _filename(self):
        return self.session.get("filename", "dataset")

    # ══════════════════════════════════════════════════════════════════════════
    # PDF — Laporan Makalah Ilmiah Lengkap
    # ══════════════════════════════════════════════════════════════════════════
    def generate_pdf(self, out_dir: str) -> str:
        import io, textwrap, matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        import matplotlib.patches as mpatches
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
            HRFlowable, PageBreak, Image, KeepTogether, ListFlowable, ListItem
        )
        from reportlab.lib.units import cm
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY, TA_RIGHT

        path = os.path.join(out_dir, f"eda_report_{self.ts_str}.pdf")

        # ── Warna ReportLab ────────────────────────────────────────────────
        C_TEAL  = colors.HexColor(TEAL)
        C_INK   = colors.HexColor(INK)
        C_INK2  = colors.HexColor(INK2)
        C_INK3  = colors.HexColor(INK3)
        C_MINT  = colors.HexColor(MINT)
        C_TEA   = colors.HexColor(TEA)
        C_GREEN = colors.HexColor(GREEN)
        C_AMBER = colors.HexColor(AMBER)
        C_RED   = colors.HexColor(RED)
        C_WHITE = colors.white
        C_GRAY  = colors.HexColor("#94a3b8")
        C_LIGHT = colors.HexColor("#e2e8f0")

        # ── Nomor halaman ──────────────────────────────────────────────────
        class NumberedCanvas:
            pass  # akan digunakan via canvasmaker

        from reportlab.pdfgen import canvas as pdfcanvas

        class PageNumCanvas(pdfcanvas.Canvas):
            def __init__(self, *args, **kwargs):
                pdfcanvas.Canvas.__init__(self, *args, **kwargs)
                self._saved_page_states = []

            def showPage(self):
                self._saved_page_states.append(dict(self.__dict__))
                self._startPage()

            def save(self):
                num_pages = len(self._saved_page_states)
                for state in self._saved_page_states:
                    self.__dict__.update(state)
                    self._draw_page_number(num_pages)
                    pdfcanvas.Canvas.showPage(self)
                pdfcanvas.Canvas.save(self)

            def _draw_page_number(self, page_count):
                pg = self._pageNumber
                if pg <= 2:
                    return  # skip cover & daftar isi
                self.setFont("Helvetica", 8)
                self.setFillColor(colors.HexColor("#94a3b8"))
                self.drawRightString(
                    A4[0] - 1.8*cm, 1.1*cm,
                    f"Halaman {pg - 2} dari {page_count - 2}"
                )
                self.setStrokeColor(colors.HexColor("#25C5E9"))
                self.setLineWidth(0.5)
                self.line(1.8*cm, 1.4*cm, A4[0]-1.8*cm, 1.4*cm)

        doc = SimpleDocTemplate(
            path, pagesize=A4,
            leftMargin=2.5*cm, rightMargin=2.5*cm,
            topMargin=2.5*cm,  bottomMargin=2.2*cm
        )

        W = A4[0] - 5.0*cm  # lebar area teks

        # ── Gaya teks ──────────────────────────────────────────────────────
        SS = getSampleStyleSheet()

        def sty(name, **kw):
            base = kw.pop("parent", SS["Normal"])
            return ParagraphStyle(name, parent=base, **kw)

        cover_title  = sty("CT",  fontSize=26, textColor=C_TEAL, fontName="Helvetica-Bold",
                            alignment=TA_CENTER, spaceAfter=6, leading=32)
        cover_sub    = sty("CS",  fontSize=12, textColor=C_TEA,  alignment=TA_CENTER,
                            spaceAfter=4, leading=16)
        cover_meta   = sty("CM",  fontSize=10, textColor=C_MINT, alignment=TA_CENTER,
                            spaceAfter=3, leading=14)
        cover_label  = sty("CL",  fontSize=9,  textColor=C_GRAY, alignment=TA_CENTER,
                            spaceAfter=2)

        h1_sty  = sty("H1",  fontSize=14, textColor=C_TEAL, fontName="Helvetica-Bold",
                       spaceBefore=18, spaceAfter=8, borderPadding=(0,0,4,0))
        h2_sty  = sty("H2",  fontSize=11, textColor=C_TEAL, fontName="Helvetica-Bold",
                       spaceBefore=12, spaceAfter=6)
        h3_sty  = sty("H3",  fontSize=10, textColor=C_TEA,  fontName="Helvetica-Bold",
                       spaceBefore=8,  spaceAfter=4)
        body_sty = sty("BD",  fontSize=9.5, textColor=C_INK,
                        leading=15, spaceAfter=6, alignment=TA_JUSTIFY)
        small_sty = sty("SM", fontSize=8,  textColor=C_GRAY, leading=12, spaceAfter=4)
        toc_sty   = sty("TC", fontSize=9.5, textColor=C_MINT, leading=14, spaceAfter=3)
        toc_h_sty = sty("TH", fontSize=11, textColor=C_TEAL, fontName="Helvetica-Bold",
                         leading=16, spaceAfter=8, alignment=TA_CENTER)
        caption_sty = sty("CAP", fontSize=8, textColor=C_GRAY, alignment=TA_CENTER,
                           spaceAfter=6, spaceBefore=2)
        bullet_sty  = sty("BUL", fontSize=9.5, textColor=C_INK,
                           leading=14, spaceAfter=3, leftIndent=14)
        quote_sty   = sty("QT",  fontSize=9,  textColor=C_TEA, leading=13,
                            leftIndent=20, rightIndent=10, borderPadding=6,
                            backColor=colors.HexColor("#EFF6FF"), spaceAfter=8,
                            spaceBefore=4)

        # ── Helper tabel ───────────────────────────────────────────────────
        def tbl_style_base(header_bg=C_TEAL, alt1=C_INK2, alt2=C_INK3):
                return TableStyle([
                    ("BACKGROUND",    (0,0), (-1, 0), header_bg),
                    ("TEXTCOLOR",     (0,0), (-1, 0), colors.white),   # teks header putih di atas teal
                    ("FONTNAME",      (0,0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE",      (0,0), (-1,-1), 7.5),
                    ("FONTNAME",      (0,1), (-1,-1), "Helvetica"),
                    ("TEXTCOLOR",     (0,1), (-1,-1), C_INK),          # teks isi gelap
                    ("ROWBACKGROUNDS",(0,1), (-1,-1), [alt1, alt2]),
                    ("GRID",          (0,0), (-1,-1), 0.3, colors.HexColor("#CBD5E1")),
                    ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
                    ("TOPPADDING",    (0,0), (-1,-1), 4),
                    ("BOTTOMPADDING", (0,0), (-1,-1), 4),
                    ("LEFTPADDING",   (0,0), (-1,-1), 6),
                    ("RIGHTPADDING",  (0,0), (-1,-1), 6),
            ])

        def make_table(rows, col_widths=None, style=None):
            t = Table(rows, colWidths=col_widths, repeatRows=1)
            t.setStyle(style or tbl_style_base())
            return t

        # ── Helper chart matplotlib → ReportLab Image ──────────────────────
        def fig_to_rl(fig, width_cm=15):
            buf = io.BytesIO()
            fig.savefig(buf, format="png", dpi=150, bbox_inches="tight",
                        facecolor=fig.get_facecolor())
            buf.seek(0)
            plt.close(fig)
            ratio = fig.get_figheight() / fig.get_figwidth()
            w = width_cm * cm
            return Image(buf, width=w, height=w * ratio)

        BG  = "#FFFFFF"
        FG  = "#1A1A1A"
        TC  = TEAL

        def ax_style(ax, title="", xlabel="", ylabel=""):
            ax.set_facecolor("#FFFFFF")
            ax.tick_params(colors=FG, labelsize=7)
            ax.xaxis.label.set_color(FG); ax.yaxis.label.set_color(FG)
            ax.title.set_color(TC)
            for spine in ax.spines.values():
                spine.set_edgecolor("#CBD5E1")
            if title:  ax.set_title(title,  fontsize=9, fontweight="bold", pad=6)
            if xlabel: ax.set_xlabel(xlabel, fontsize=7)
            if ylabel: ax.set_ylabel(ylabel, fontsize=7)
            ax.grid(True, color="#CBD5E1", alpha=0.5, linewidth=0.5)

        # ══════════════════════════════════════════════════════════════════
        story = []
        num_stats = self._num_stats()
        cat_stats = self._cat_stats()
        insights  = self._insights()
        ts        = self._ts()
        fname     = self._filename()
        n_rows    = len(self.df)
        n_cols    = len(self.df.columns)
        num_cols  = self.df.select_dtypes(include="number").columns.tolist()
        cat_cols  = self.df.select_dtypes(include=["object","category"]).columns.tolist()

        # ─────────────────────────────────────────────────────────────────
        # HALAMAN 1 — COVER
        # ─────────────────────────────────────────────────────────────────
        story.append(Spacer(1, 1.8*cm))

        # Garis dekorasi atas
        story.append(HRFlowable(width="100%", thickness=3, color=C_TEAL, spaceAfter=16))
        story.append(Paragraph("LAPORAN EKSPLORASI DATA OTOMATIS", cover_label))
        story.append(Spacer(1, 0.3*cm))
        story.append(Paragraph("Auto EDA Insight", cover_title))
        story.append(Spacer(1, 0.2*cm))
        story.append(Paragraph("Full-Dashboard Analytical Report", cover_sub))
        story.append(HRFlowable(width="60%", thickness=1, color=C_TEAL, spaceAfter=20))

        story.append(Spacer(1, 0.8*cm))

        # Kotak informasi dataset
        cover_rows = [
            ["Dataset", fname],
            ["Jumlah Baris", f"{n_rows:,} baris"],
            ["Jumlah Kolom", f"{n_cols} kolom"],
            ["Variabel Numerik", f"{len(num_cols)} kolom"],
            ["Variabel Kategorik", f"{len(cat_cols)} kolom"],
            ["Tanggal Dibuat", self.now],
        ]
        cover_tbl = Table(
            [["Properti", "Nilai"]] + cover_rows,
            colWidths=[5.5*cm, W - 5.5*cm]
        )
        cover_tbl.setStyle(TableStyle([
            ("BACKGROUND",    (0,0), (-1, 0), C_TEAL),
            ("TEXTCOLOR",     (0,0), (-1, 0), C_INK),
            ("FONTNAME",      (0,0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",      (0,0), (-1,-1), 9),
            ("FONTNAME",      (0,1), (-1,-1), "Helvetica"),
            ("TEXTCOLOR",     (0,0), (0,-1),  C_TEAL),
            ("TEXTCOLOR",     (1,1), (-1,-1), C_MINT),
            ("FONTNAME",      (0,1), (0,-1),  "Helvetica-Bold"),
            ("ROWBACKGROUNDS",(0,1), (-1,-1), [C_INK2, C_INK3]),
            ("GRID",          (0,0), (-1,-1), 0.4, colors.HexColor("#238689")),
            ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
            ("TOPPADDING",    (0,0), (-1,-1), 7),
            ("BOTTOMPADDING", (0,0), (-1,-1), 7),
            ("LEFTPADDING",   (0,0), (-1,-1), 12),
        ]))
        story.append(cover_tbl)

        story.append(Spacer(1, 1.5*cm))
        story.append(HRFlowable(width="100%", thickness=3, color=C_TEAL, spaceAfter=12))
        story.append(Paragraph("SD-1306 Data Science Programming", cover_meta))
        story.append(Paragraph("Institut Teknologi Sains Bandung", cover_meta))
        story.append(Spacer(1, 0.2*cm))
        story.append(Paragraph(f"Laporan ini digenerate secara otomatis oleh sistem Auto EDA Insight pada {self.now}.",
                                cover_label))
        story.append(PageBreak())

        # ─────────────────────────────────────────────────────────────────
        # HALAMAN 2 — DAFTAR ISI
        # ─────────────────────────────────────────────────────────────────
        story.append(Spacer(1, 0.4*cm))
        story.append(Paragraph("DAFTAR ISI", toc_h_sty))
        story.append(HRFlowable(width="100%", thickness=1.5, color=C_TEAL, spaceAfter=16))

        toc_entries = [
            ("I",   "Pendahuluan",                              "3"),
            ("II",  "Tinjauan Dataset",                         "3"),
            ("III", "Analisis Statistik Deskriptif",            "4"),
            ("",    "  A. Variabel Numerik",                    "4"),
            ("",    "  B. Variabel Kategorik",                  "5"),
            ("IV",  "Visualisasi Data",                         "6"),
            ("",    "  A. Distribusi Variabel Numerik",         "6"),
            ("",    "  B. Distribusi Variabel Kategorik",       "7"),
            ("",    "  C. Korelasi Antar Variabel",             "7"),
            ("V",   "Analisis Deret Waktu",                     "8"),
            ("VI",  "Temuan dan Auto Insights",                 "8"),
            ("VII", "Interpretasi",                             "9"),
            ("VIII","Kesimpulan",                               "10"),
            ("IX",  "Saran",                                    "10"),
            ("X",   "Lampiran — Pratinjau Data",                "11"),
        ]
        for roman, title, pg in toc_entries:
            if roman:
                txt = f"<b>Bab {roman}</b> — {title}"
                s = toc_sty
            else:
                txt = title
                s = small_sty
            dot_space = "." * max(2, 60 - len(title))
            toc_row = Table(
                [[Paragraph(txt, s), Paragraph(f"<para alignment='right'>{pg}</para>", s)]],
                colWidths=[W - 1.5*cm, 1.5*cm]
            )
            toc_row.setStyle(TableStyle([
                ("VALIGN",       (0,0), (-1,-1), "BOTTOM"),
                ("BOTTOMPADDING",(0,0), (-1,-1), 2),
                ("LINEBELOW",    (0,0), (-2,-1), 0.3, colors.HexColor("#1a3a52")),
            ]))
            story.append(toc_row)
            story.append(Spacer(1, 0.05*cm))

        story.append(PageBreak())

        # ─────────────────────────────────────────────────────────────────
        # BAB I — PENDAHULUAN
        # ─────────────────────────────────────────────────────────────────
        story.append(Paragraph("Bab I — Pendahuluan", h1_sty))
        story.append(HRFlowable(width="100%", thickness=1, color=C_TEAL, spaceAfter=10))

        story.append(Paragraph(
            "Eksplorasi Data (Exploratory Data Analysis / EDA) merupakan tahapan fundamental "
            "dalam alur kerja ilmu data modern. Proses ini bertujuan untuk memahami struktur, "
            "distribusi, pola, anomali, dan hubungan antar variabel dalam sebuah dataset sebelum "
            "dilakukan pemodelan statistik atau pembelajaran mesin. Pemahaman mendalam terhadap "
            "data mentah sangat kritis untuk menghasilkan model yang akurat dan andal.",
            body_sty
        ))
        story.append(Paragraph(
            f"Laporan ini menyajikan hasil eksplorasi otomatis terhadap dataset <b>{html.escape(fname)}</b> "
            f"yang terdiri dari <b>{n_rows:,} baris observasi</b> dan <b>{n_cols} variabel</b> "
            f"({len(num_cols)} numerik, {len(cat_cols)} kategorik). Seluruh analisis, visualisasi, "
            "dan interpretasi dalam laporan ini dihasilkan secara otomatis oleh sistem <i>Auto EDA Insight</i> "
            "yang dikembangkan dalam kerangka mata kuliah SD-1306 Data Science Programming, "
            "Institut Teknologi Sains Bandung.",
            body_sty
        ))
        story.append(Paragraph(
            "Tujuan utama laporan ini adalah: (1) memberikan gambaran menyeluruh tentang karakteristik "
            "dataset; (2) mengidentifikasi isu kualitas data seperti nilai yang hilang dan pencilan; "
            "(3) mengungkap pola dan distribusi variabel; (4) mendeteksi hubungan antar variabel; serta "
            "(5) memberikan rekomendasi berbasis data untuk tahapan analisis berikutnya.",
            body_sty
        ))

        # BAB II — TINJAUAN DATASET
        story.append(Spacer(1, 0.4*cm))
        story.append(Paragraph("Bab II — Tinjauan Dataset", h1_sty))
        story.append(HRFlowable(width="100%", thickness=1, color=C_TEAL, spaceAfter=10))

        # Hitung missing
        total_missing = int(self.df.isnull().sum().sum())
        total_cells   = n_rows * n_cols
        missing_pct   = (total_missing / total_cells * 100) if total_cells > 0 else 0
        dup_rows      = int(self.df.duplicated().sum())

        story.append(Paragraph(
            f"Dataset <b>{html.escape(fname)}</b> merupakan kumpulan data yang memuat {n_rows:,} "
            f"observasi dengan {n_cols} atribut. Komposisi variabel terdiri atas "
            f"{len(num_cols)} variabel bertipe numerik (kontinue maupun diskrit) dan "
            f"{len(cat_cols)} variabel bertipe kategorik. Total sel data yang tersedia adalah "
            f"{total_cells:,} sel, di mana terdapat {total_missing:,} nilai yang hilang "
            f"({missing_pct:.2f}% dari keseluruhan data). Selain itu, terdeteksi {dup_rows:,} "
            f"baris duplikat yang perlu diperhatikan dalam tahap pra-pemrosesan.",
            body_sty
        ))

        # Tabel ringkasan dataset
        summary_rows = [
            ["Metrik", "Nilai", "Keterangan"],
            ["Total Baris",          f"{n_rows:,}",        "Jumlah observasi"],
            ["Total Kolom",          f"{n_cols}",           "Jumlah variabel"],
            ["Variabel Numerik",     f"{len(num_cols)}",    ", ".join(num_cols[:5]) + ("..." if len(num_cols)>5 else "")],
            ["Variabel Kategorik",   f"{len(cat_cols)}",    ", ".join(cat_cols[:5]) + ("..." if len(cat_cols)>5 else "")],
            ["Total Nilai Hilang",   f"{total_missing:,}",  f"{missing_pct:.2f}% dari total sel"],
            ["Baris Duplikat",       f"{dup_rows:,}",       "Perlu ditangani sebelum modeling"],
            ["Memori (estimasi)",    f"{self.df.memory_usage(deep=True).sum()/1024:.1f} KB", "Ukuran di memori"],
        ]
        cw3 = [4*cm, 3*cm, W-7*cm]
        story.append(make_table(summary_rows, col_widths=cw3))
        story.append(Paragraph("Tabel 1. Ringkasan Umum Dataset", caption_sty))

        if missing_pct > 10:
            story.append(Paragraph(
                f"⚠  Persentase nilai hilang cukup signifikan ({missing_pct:.1f}%). "
                "Disarankan untuk menerapkan strategi imputasi (mean/median/modus/KNN) "
                "atau menghapus baris/kolom dengan missing value tinggi sebelum analisis lanjutan.",
                quote_sty
            ))

        story.append(PageBreak())

        # ─────────────────────────────────────────────────────────────────
        # BAB III — STATISTIK DESKRIPTIF
        # ─────────────────────────────────────────────────────────────────
        story.append(Paragraph("Bab III — Analisis Statistik Deskriptif", h1_sty))
        story.append(HRFlowable(width="100%", thickness=1, color=C_TEAL, spaceAfter=10))
        story.append(Paragraph(
            "Statistik deskriptif memberikan ringkasan kuantitatif mengenai karakteristik dasar "
            "setiap variabel dalam dataset. Analisis ini mencakup ukuran pemusatan (mean, median, modus), "
            "ukuran penyebaran (standar deviasi, varians, IQR), ukuran bentuk distribusi (skewness, kurtosis), "
            "serta informasi mengenai nilai yang hilang dan pencilan.",
            body_sty
        ))

        # A. Numerik
        story.append(Paragraph("A. Variabel Numerik", h2_sty))
        if num_stats:
            story.append(Paragraph(
                f"Terdapat {len(num_stats)} variabel numerik yang dianalisis. "
                "Tabel berikut menyajikan statistik deskriptif lengkap untuk masing-masing variabel, "
                "termasuk hasil uji normalitas menggunakan Shapiro-Wilk dan deteksi pencilan dengan metode IQR.",
                body_sty
            ))
            hdr_n = ["Kolom","N","Mean","Median","Std","Min","Max","Skew","Kurt","Missing","Normal","Outlier"]
            keys_n = ["column","count","mean","median","std","min","max","skewness","kurtosis","missing_count","normality","outliers"]
            rows_n = [hdr_n]
            for s in num_stats:
                def fmt(v, d=3):
                    try: return f"{float(v):.{d}f}"
                    except: return str(v) if v not in (None,"") else "—"
                row = [
                    str(s.get("column","—")),
                    str(s.get("count","—")),
                    fmt(s.get("mean")),
                    fmt(s.get("median")),
                    fmt(s.get("std")),
                    fmt(s.get("min")),
                    fmt(s.get("max")),
                    fmt(s.get("skewness",0), 2),
                    fmt(s.get("kurtosis",0), 2),
                    str(s.get("missing_count","0")),
                    "✓ Normal" if s.get("normality")=="Normal" else "✗ Tidak",
                    str(s.get("outliers","0")),
                ]
                rows_n.append(row)

            cw_n = [3.2*cm] + [1.5*cm]*11
            ts_n = tbl_style_base()
            for i, s in enumerate(num_stats, start=1):
                bg = C_GREEN if s.get("normality")=="Normal" else C_RED
                ts_n.add("BACKGROUND", (10,i), (10,i), bg)
                ts_n.add("TEXTCOLOR",  (10,i), (10,i), C_INK)
                if int(s.get("outliers",0) or 0) > 0:
                    ts_n.add("BACKGROUND", (11,i), (11,i), C_AMBER)
                    ts_n.add("TEXTCOLOR",  (11,i), (11,i), C_INK)
            story.append(make_table(rows_n, col_widths=cw_n, style=ts_n))
            story.append(Paragraph("Tabel 2. Statistik Deskriptif Variabel Numerik", caption_sty))

            # Interpretasi paragraf otomatis
            normal_cols  = [s["column"] for s in num_stats if s.get("normality")=="Normal"]
            skewed_cols  = [s["column"] for s in num_stats if abs(float(s.get("skewness",0) or 0)) > 1]
            outlier_cols = [s["column"] for s in num_stats if int(s.get("outliers",0) or 0) > 0]

            if normal_cols:
                story.append(Paragraph(
                    f"Berdasarkan uji normalitas, variabel <b>{html.escape(', '.join(normal_cols))}</b> "
                    "memenuhi asumsi distribusi normal (p-value > 0.05), sehingga dapat "
                    "dianalisis menggunakan metode statistik parametrik.",
                    body_sty
                ))
            if skewed_cols:
                story.append(Paragraph(
                    f"Variabel <b>{html.escape(', '.join(skewed_cols))}</b> menunjukkan distribusi yang condong "
                    "(skewness |s| > 1), mengindikasikan adanya ketidaksimetrisan yang perlu "
                    "dipertimbangkan, misalnya melalui transformasi logaritmik atau Box-Cox.",
                    body_sty
                ))
            if outlier_cols:
                story.append(Paragraph(
                    f"Variabel <b>{html.escape(', '.join(outlier_cols))}</b> terdeteksi memiliki pencilan "
                    "berdasarkan metode IQR (nilai di luar rentang Q1−1.5×IQR hingga Q3+1.5×IQR). "
                    "Pencilan ini dapat berdampak signifikan terhadap hasil model prediktif "
                    "dan perlu ditangani dengan capping, winsorizing, atau penghapusan.",
                    body_sty
                ))
        else:
            story.append(Paragraph("Tidak ditemukan variabel numerik dalam dataset.", body_sty))

        story.append(Spacer(1, 0.3*cm))

        # B. Kategorik
        story.append(Paragraph("B. Variabel Kategorik", h2_sty))
        if cat_stats:
            story.append(Paragraph(
                f"Terdapat {len(cat_stats)} variabel kategorik dalam dataset. "
                "Analisis meliputi jumlah kategori unik, nilai modus beserta frekuensinya, "
                "serta persentase nilai yang hilang untuk setiap variabel.",
                body_sty
            ))
            hdr_c = ["Kolom","Jumlah","Unik","Modus","Freq Modus","% Modus","Missing","% Missing"]
            keys_c = ["column","count","unique","mode","mode_freq","mode_pct","missing_count","missing_pct"]
            rows_c = [hdr_c] + [[str(s.get(k,"—")) for k in keys_c] for s in cat_stats]
            cw_c = [3*cm, 1.8*cm, 1.5*cm, 3.5*cm, 2.2*cm, 2*cm, 1.8*cm, 2.2*cm]
            story.append(make_table(rows_c, col_widths=cw_c))
            story.append(Paragraph("Tabel 3. Statistik Deskriptif Variabel Kategorik", caption_sty))

            high_card = [s["column"] for s in cat_stats if int(s.get("unique",0) or 0) > 20]
            if high_card:
                story.append(Paragraph(
                    f"Variabel <b>{html.escape(', '.join(high_card))}</b> memiliki kardinalitas tinggi (>20 kategori unik). "
                    "Variabel semacam ini berpotensi menimbulkan masalah curse of dimensionality jika "
                    "digunakan langsung dalam model. Pertimbangkan encoding yang tepat seperti "
                    "target encoding atau feature hashing.",
                    body_sty
                ))
        else:
            story.append(Paragraph("Tidak ditemukan variabel kategorik dalam dataset.", body_sty))

        story.append(PageBreak())

        # ─────────────────────────────────────────────────────────────────
        # BAB IV — VISUALISASI DATA
        # ─────────────────────────────────────────────────────────────────
        story.append(Paragraph("Bab IV — Visualisasi Data", h1_sty))
        story.append(HRFlowable(width="100%", thickness=1, color=C_TEAL, spaceAfter=10))
        story.append(Paragraph(
            "Visualisasi data merupakan alat yang sangat efektif untuk memahami pola dan "
            "distribusi data secara intuitif. Bagian ini menyajikan berbagai jenis grafik "
            "yang dihasilkan secara otomatis dari dataset, mencakup histogram distribusi, "
            "grafik batang kategorik, dan heatmap korelasi.",
            body_sty
        ))

        # A. Distribusi Numerik
        story.append(Paragraph("A. Distribusi Variabel Numerik", h2_sty))
        num_plot_cols = num_cols[:6]  # maks 6
        if num_plot_cols:
            story.append(Paragraph(
                f"Grafik berikut menampilkan histogram dan kurva distribusi untuk "
                f"{len(num_plot_cols)} variabel numerik pertama dalam dataset. "
                "Garis merah putus-putus menunjukkan nilai mean, sedangkan garis hijau menunjukkan median.",
                body_sty
            ))
            ncols_fig = min(3, len(num_plot_cols))
            nrows_fig = (len(num_plot_cols) + ncols_fig - 1) // ncols_fig
            fig, axes = plt.subplots(nrows_fig, ncols_fig,
                                     figsize=(5*ncols_fig, 3.2*nrows_fig),
                                     facecolor=BG)
            axes_flat = np.array(axes).flatten()
            for i, col in enumerate(num_plot_cols):
                ax = axes_flat[i]
                data = self.df[col].dropna()
                ax.hist(data, bins=30, color=TEAL, alpha=0.75, edgecolor=INK3, linewidth=0.4)
                ax.axvline(data.mean(),   color=RED,   linestyle="--", linewidth=1.2, label="Mean")
                ax.axvline(data.median(), color=GREEN, linestyle="--", linewidth=1.2, label="Median")
                ax_style(ax, title=col, xlabel="Nilai", ylabel="Frekuensi")
                ax.legend(fontsize=6, facecolor=INK2, labelcolor=FG, framealpha=0.7)
            for j in range(len(num_plot_cols), len(axes_flat)):
                axes_flat[j].set_visible(False)
            plt.tight_layout(pad=0.8)
            story.append(fig_to_rl(fig, width_cm=16))
            story.append(Paragraph("Gambar 1. Distribusi Variabel Numerik (Histogram)", caption_sty))
        else:
            story.append(Paragraph("Tidak tersedia variabel numerik untuk divisualisasikan.", body_sty))

        story.append(Spacer(1, 0.3*cm))

        # B. Distribusi Kategorik
        story.append(Paragraph("B. Distribusi Variabel Kategorik", h2_sty))
        cat_plot_cols = cat_cols[:4]
        if cat_plot_cols:
            story.append(Paragraph(
                f"Grafik batang horizontal berikut menampilkan distribusi frekuensi untuk "
                f"{len(cat_plot_cols)} variabel kategorik. Hanya ditampilkan 10 kategori "
                "teratas berdasarkan frekuensi kemunculannya.",
                body_sty
            ))
            ncols_c = min(2, len(cat_plot_cols))
            nrows_c = (len(cat_plot_cols) + ncols_c - 1) // ncols_c
            fig2, axes2 = plt.subplots(nrows_c, ncols_c,
                                       figsize=(6*ncols_c, 3.5*nrows_c),
                                       facecolor=BG)
            axes2_flat = np.array(axes2).flatten()
            colors_bar = [TEAL, "#22c55e", "#f59e0b", "#a78bfa"]
            for i, col in enumerate(cat_plot_cols):
                ax2 = axes2_flat[i]
                vc = self.df[col].value_counts().head(10)
                bars = ax2.barh(range(len(vc)), vc.values,
                                color=colors_bar[i % len(colors_bar)], alpha=0.8)
                ax2.set_yticks(range(len(vc)))
                ax2.set_yticklabels([str(v)[:18] for v in vc.index], fontsize=7)
                for bar, val in zip(bars, vc.values):
                    ax2.text(bar.get_width()*1.01, bar.get_y()+bar.get_height()/2,
                             f"{val:,}", va="center", fontsize=6, color=FG)
                ax_style(ax2, title=col, xlabel="Frekuensi")
            for j in range(len(cat_plot_cols), len(axes2_flat)):
                axes2_flat[j].set_visible(False)
            plt.tight_layout(pad=0.8)
            story.append(fig_to_rl(fig2, width_cm=16))
            story.append(Paragraph("Gambar 2. Distribusi Frekuensi Variabel Kategorik", caption_sty))
        else:
            story.append(Paragraph("Tidak tersedia variabel kategorik untuk divisualisasikan.", body_sty))

        story.append(Spacer(1, 0.3*cm))

        # C. Korelasi
        story.append(Paragraph("C. Matriks Korelasi Antar Variabel Numerik", h2_sty))
        if len(num_cols) >= 2:
            story.append(Paragraph(
                "Heatmap korelasi Pearson berikut menggambarkan kekuatan dan arah hubungan "
                "linear antar variabel numerik. Nilai mendekati +1 menunjukkan korelasi positif kuat, "
                "mendekati −1 korelasi negatif kuat, dan mendekati 0 berarti tidak ada korelasi linear.",
                body_sty
            ))
            corr_cols = num_cols[:10]
            corr_df   = self.df[corr_cols].corr()
            fig3, ax3 = plt.subplots(figsize=(max(5, len(corr_cols)*0.9),
                                               max(4, len(corr_cols)*0.8)),
                                     facecolor=BG)
            mask_upper = np.zeros_like(corr_df, dtype=bool)
            mask_upper[np.triu_indices_from(mask_upper, k=1)] = True
            data_plot = corr_df.values.copy()
            data_plot[mask_upper] = np.nan

            from matplotlib.colors import LinearSegmentedColormap
            cmap = LinearSegmentedColormap.from_list("eda",
                [RED, INK2, TEAL], N=256)
            im = ax3.imshow(data_plot, cmap=cmap, vmin=-1, vmax=1, aspect="auto")
            ax3.set_xticks(range(len(corr_cols)))
            ax3.set_yticks(range(len(corr_cols)))
            ax3.set_xticklabels([c[:12] for c in corr_cols], rotation=45, ha="right", fontsize=7, color=FG)
            ax3.set_yticklabels([c[:12] for c in corr_cols], fontsize=7, color=FG)
            for i in range(len(corr_cols)):
                for j in range(len(corr_cols)):
                    if not mask_upper[i,j] and not np.isnan(data_plot[i,j]):
                        val = data_plot[i,j]
                        ax3.text(j, i, f"{val:.2f}", ha="center", va="center",
                                 fontsize=6.5, color="white" if abs(val)>0.4 else FG,
                                 fontweight="bold")
            cb = plt.colorbar(im, ax=ax3, fraction=0.04, pad=0.02)
            cb.ax.tick_params(labelsize=7, colors=FG)
            cb.outline.set_edgecolor("#238689")
            ax3.set_facecolor(INK2)
            for spine in ax3.spines.values():
                spine.set_edgecolor("#238689")
            ax3.set_title("Matriks Korelasi Pearson", color=TEAL, fontsize=10, fontweight="bold", pad=10)
            plt.tight_layout()
            story.append(fig_to_rl(fig3, width_cm=14))
            story.append(Paragraph("Gambar 3. Heatmap Matriks Korelasi Pearson", caption_sty))

            # Temuan korelasi tinggi
            high_corr = []
            for i in range(len(corr_cols)):
                for j in range(i+1, len(corr_cols)):
                    v = corr_df.iloc[i, j]
                    if abs(v) >= 0.7:
                        high_corr.append((corr_cols[i], corr_cols[j], v))
            if high_corr:
                story.append(Paragraph(
                    "Pasangan variabel dengan korelasi tinggi (|r| ≥ 0.7) terdeteksi: " +
                    "; ".join([f"<b>{html.escape(str(a))}</b> & <b>{html.escape(str(b))}</b> (r={v:.2f})" for a,b,v in high_corr[:5]]) +
                    ". Korelasi tinggi ini berpotensi menimbulkan multikolinearitas dalam model regresi.",
                    body_sty
                ))
        else:
            story.append(Paragraph(
                "Matriks korelasi tidak dapat dibuat karena kurang dari 2 variabel numerik.", body_sty
            ))

        story.append(PageBreak())

        # ─────────────────────────────────────────────────────────────────
        # BAB V — ANALISIS DERET WAKTU
        # ─────────────────────────────────────────────────────────────────
        story.append(Paragraph("Bab V — Analisis Deret Waktu", h1_sty))
        story.append(HRFlowable(width="100%", thickness=1, color=C_TEAL, spaceAfter=10))

        if ts.get("has_timeseries"):
            dt_col  = ts.get("date_column","")
            ts_num  = num_cols[:3]
            story.append(Paragraph(
                f"Dataset ini memiliki dimensi waktu dengan kolom tanggal <b>{html.escape(str(dt_col))}</b>. "
                "Analisis deret waktu dilakukan untuk mengidentifikasi tren jangka panjang, "
                "pola musiman, serta fluktuasi variabel numerik sepanjang periode observasi.",
                body_sty
            ))

            if dt_col in self.df.columns and ts_num:
                df_s = self.df.sort_values(dt_col).copy()

                # Plot time series
                fig_ts, ax_ts = plt.subplots(figsize=(14, 4), facecolor=BG)
                colors_ts = [TEAL, GREEN, AMBER]
                for ci, c in enumerate(ts_num):
                    s = df_s[c].dropna()
                    if len(s) < 2: continue
                    idx = range(len(s))
                    ax_ts.plot(idx, s.values, color=colors_ts[ci % 3],
                               linewidth=1, alpha=0.8, label=c)
                    # rolling mean
                    window = max(3, len(s)//20)
                    rm = s.rolling(window, min_periods=1).mean()
                    ax_ts.plot(idx, rm.values, color=colors_ts[ci % 3],
                               linewidth=2, linestyle="--", alpha=0.5)
                ax_style(ax_ts, title="Tren Deret Waktu dengan Moving Average",
                         xlabel="Indeks Waktu", ylabel="Nilai")
                ax_ts.legend(fontsize=7, facecolor=INK2, labelcolor=FG, framealpha=0.7)
                plt.tight_layout()
                story.append(fig_to_rl(fig_ts, width_cm=16))
                story.append(Paragraph("Gambar 4. Grafik Deret Waktu dan Moving Average", caption_sty))

                # Tabel ringkasan TS
                ts_hdr  = ["Variabel","Tgl Awal","Tgl Akhir","Min","Max","Mean","Std","Tren"]
                ts_rows = [ts_hdr]
                for c in ts_num:
                    s = df_s[c].dropna()
                    if len(s) < 2: continue
                    fh = s.iloc[:len(s)//2].mean()
                    sh = s.iloc[len(s)//2:].mean()
                    tren = "↑ Naik" if sh > fh else "↓ Turun"
                    fd = str(df_s[dt_col].dropna().iloc[0])[:10]
                    ld = str(df_s[dt_col].dropna().iloc[-1])[:10]
                    ts_rows.append([c, fd, ld,
                                    f"{s.min():.2f}", f"{s.max():.2f}",
                                    f"{s.mean():.2f}", f"{s.std():.2f}", tren])
                cw_ts_t = [3*cm, 2.5*cm, 2.5*cm, 1.8*cm, 1.8*cm, 1.8*cm, 1.8*cm, 2.8*cm]
                ts_style = tbl_style_base()
                for i, c in enumerate(ts_num, 1):
                    if i < len(ts_rows):
                        up = "↑" in ts_rows[i][-1]
                        ts_style.add("BACKGROUND", (-1,i), (-1,i), C_GREEN if up else C_RED)
                        ts_style.add("TEXTCOLOR",  (-1,i), (-1,i), C_INK)
                story.append(make_table(ts_rows, col_widths=cw_ts_t, style=ts_style))
                story.append(Paragraph("Tabel 4. Ringkasan Statistik Deret Waktu", caption_sty))

                # Interpretasi
                for c in ts_num:
                    s = df_s[c].dropna()
                    if len(s) < 2: continue
                    pct_change = ((s.iloc[-1] - s.iloc[0]) / (abs(s.iloc[0]) or 1)) * 100
                    story.append(Paragraph(
                        f"Variabel <b>{html.escape(str(c))}</b> mengalami perubahan sebesar <b>{pct_change:+.1f}%</b> "
                        f"dari nilai awal {s.iloc[0]:.2f} hingga nilai akhir {s.iloc[-1]:.2f}. "
                        f"Rata-rata nilai adalah {s.mean():.2f} dengan standar deviasi {s.std():.2f}, "
                        f"yang mengindikasikan {'volatilitas tinggi' if s.std()/s.mean() > 0.3 else 'variasi moderat'} "
                        "sepanjang periode observasi.",
                        body_sty
                    ))
        else:
            story.append(Paragraph(
                "Dataset ini tidak memiliki kolom bertipe datetime yang terdeteksi secara otomatis. "
                "Analisis deret waktu tidak dapat dilakukan. Jika dataset memiliki dimensi temporal "
                "dalam format string, disarankan untuk mengonversinya ke format datetime terlebih dahulu "
                "menggunakan pd.to_datetime() sebelum analisis lebih lanjut.",
                body_sty
            ))

        story.append(PageBreak())

        # ─────────────────────────────────────────────────────────────────
        # BAB VI — TEMUAN DAN AUTO INSIGHTS
        # ─────────────────────────────────────────────────────────────────
        story.append(Paragraph("Bab VI — Temuan dan Auto Insights", h1_sty))
        story.append(HRFlowable(width="100%", thickness=1, color=C_TEAL, spaceAfter=10))
        story.append(Paragraph(
            "Sistem Auto EDA Insight secara otomatis mendeteksi pola, anomali, dan temuan "
            "penting dalam dataset menggunakan serangkaian algoritma heuristik berbasis "
            "statistik. Berikut adalah daftar insight yang berhasil diidentifikasi.",
            body_sty
        ))

        if insights:
            # Klasifikasi per tipe
            by_type = {"success": [], "warning": [], "danger": [], "info": []}
            for ins in insights:
                tp = ins.get("type","info")
                by_type.setdefault(tp, []).append(ins)

            type_labels = {
                "danger":  ("⚠ Kritis",    RED),
                "warning": ("⚡ Perhatian", AMBER),
                "success": ("✓ Positif",   GREEN),
                "info":    ("ℹ Informasi", TEAL),
            }
            for tp_key in ["danger","warning","success","info"]:
                grp = by_type.get(tp_key, [])
                if not grp: continue
                label, color = type_labels[tp_key]
                story.append(Paragraph(f"{label} ({len(grp)} temuan)", h3_sty))
                for ins in grp:
                    text = ins.get("text","").replace("<b>","").replace("</b>","")
                    title = ins.get("title","")
                    sym   = insight_symbol(ins)
                    story.append(Paragraph(
                        f"[{sym}] <b>{html.escape(title)}</b><br/>{html.escape(text)}", bullet_sty
                    ))
                story.append(Spacer(1, 0.2*cm))

            # Tabel ringkasan insights
            ins_hdr = ["#","Tipe","Judul","Keterangan"]
            ins_rows = [ins_hdr]
            for i, ins in enumerate(insights, 1):
                text  = ins.get("text","").replace("<b>","").replace("</b>","")
                ins_rows.append([
                    str(i),
                    ins.get("type","info").upper(),
                    ins.get("title","")[:40],
                    text[:120]
                ])
            cw_ins = [0.8*cm, 2.2*cm, 4.5*cm, W-7.5*cm]
            ts_ins = tbl_style_base()
            for i, ins in enumerate(insights, 1):
                tp = ins.get("type","info")
                bg = {"success":C_GREEN,"warning":C_AMBER,"danger":C_RED}.get(tp, C_TEAL)
                fg = C_INK if tp in ("success","warning") else C_MINT
                ts_ins.add("BACKGROUND",(1,i),(1,i), bg)
                ts_ins.add("TEXTCOLOR", (1,i),(1,i), fg)
                ts_ins.add("FONTNAME",  (1,i),(1,i), "Helvetica-Bold")
            story.append(make_table(ins_rows, col_widths=cw_ins, style=ts_ins))
            story.append(Paragraph("Tabel 5. Ringkasan Auto Insights", caption_sty))
        else:
            story.append(Paragraph(
                "Tidak ada insight yang digenerate. Pastikan analisis statistik telah "
                "dijalankan sebelum ekspor laporan.", body_sty
            ))

        story.append(PageBreak())

        # ─────────────────────────────────────────────────────────────────
        # BAB VII — INTERPRETASI
        # ─────────────────────────────────────────────────────────────────
        story.append(Paragraph("Bab VII — Interpretasi", h1_sty))
        story.append(HRFlowable(width="100%", thickness=1, color=C_TEAL, spaceAfter=10))

        story.append(Paragraph(
            "Interpretasi data merupakan proses pemberian makna terhadap hasil analisis statistik "
            "dan visual yang telah dilakukan. Bagian ini menyajikan pemahaman mendalam terhadap "
            "karakteristik dataset berdasarkan temuan-temuan yang diperoleh.",
            body_sty
        ))

        story.append(Paragraph("Kualitas Data", h2_sty))
        if total_missing == 0:
            story.append(Paragraph(
                "Dataset ini memiliki kualitas yang sangat baik dari sisi kelengkapan data — "
                "tidak ditemukan satu pun nilai yang hilang. Hal ini menjadi modal kuat untuk "
                "analisis lanjutan tanpa perlu tahap imputasi data.",
                body_sty
            ))
        elif missing_pct < 5:
            story.append(Paragraph(
                f"Persentase nilai yang hilang tergolong rendah ({missing_pct:.1f}%), "
                "sehingga dampaknya terhadap analisis relatif minimal. Imputasi sederhana "
                "menggunakan mean atau median sudah cukup memadai.",
                body_sty
            ))
        else:
            story.append(Paragraph(
                f"Persentase nilai yang hilang cukup signifikan ({missing_pct:.1f}%). "
                "Strategi penanganan yang tepat sangat dibutuhkan agar kualitas analisis "
                "tetap terjaga. Pilihan penanganan meliputi: imputasi KNN, MICE, atau "
                "penghapusan variabel jika >40% nilainya hilang.",
                body_sty
            ))

        if dup_rows > 0:
            story.append(Paragraph(
                f"Ditemukan {dup_rows:,} baris duplikat ({dup_rows/n_rows*100:.1f}% dari total data). "
                "Duplikasi dapat menimbulkan bias dalam estimasi parameter statistik dan performa model. "
                "Disarankan untuk menghapus baris duplikat sebelum melanjutkan ke tahap pemodelan.",
                body_sty
            ))

        story.append(Paragraph("Distribusi dan Normalitas", h2_sty))
        if num_stats:
            n_normal    = sum(1 for s in num_stats if s.get("normality")=="Normal")
            n_notnormal = len(num_stats) - n_normal
            story.append(Paragraph(
                f"Dari {len(num_stats)} variabel numerik, {n_normal} variabel mengikuti "
                f"distribusi normal dan {n_notnormal} variabel tidak normal berdasarkan uji Shapiro-Wilk. "
                "Untuk variabel yang tidak normal, metode non-parametrik lebih sesuai digunakan, "
                "atau dapat dilakukan transformasi data terlebih dahulu (log, sqrt, Box-Cox) "
                "untuk mendekati distribusi normal sebelum penerapan model parametrik.",
                body_sty
            ))

        story.append(Paragraph("Pencilan (Outliers)", h2_sty))
        if num_stats:
            total_outliers = sum(int(s.get("outliers",0) or 0) for s in num_stats)
            story.append(Paragraph(
                f"Total {total_outliers:,} pencilan terdeteksi menggunakan metode IQR. "
                "Pencilan dapat merupakan kesalahan pengukuran (error) atau observasi yang "
                "memang ekstrem secara alami (genuinely extreme values). Analisis domain "
                "diperlukan untuk membedakan keduanya sebelum memutuskan penanganan. "
                "Metode yang umum digunakan antara lain: capping pada persentil ke-1 dan ke-99, "
                "winsorizing, atau penggunaan algoritma yang robust terhadap outlier.",
                body_sty
            ))

        # ─────────────────────────────────────────────────────────────────
        # BAB VIII — KESIMPULAN
        # ─────────────────────────────────────────────────────────────────
        story.append(Spacer(1, 0.3*cm))
        story.append(Paragraph("Bab VIII — Kesimpulan", h1_sty))
        story.append(HRFlowable(width="100%", thickness=1, color=C_TEAL, spaceAfter=10))

        story.append(Paragraph(
            f"Analisis eksplorasi data terhadap dataset <b>{html.escape(fname)}</b> "
            f"yang terdiri dari {n_rows:,} baris dan {n_cols} kolom telah berhasil dilakukan "
            "secara otomatis. Berikut adalah poin-poin kesimpulan utama dari analisis ini:",
            body_sty
        ))

        conclusions = [
            f"Dataset memuat {n_rows:,} observasi dengan {n_cols} variabel ({len(num_cols)} numerik, "
            f"{len(cat_cols)} kategorik), memberikan cakupan data yang {'memadai' if n_rows > 1000 else 'terbatas'}.",

            f"Kualitas data {'sangat baik' if missing_pct < 1 else 'memerlukan perhatian'} dengan "
            f"persentase nilai hilang sebesar {missing_pct:.1f}% dan {dup_rows:,} baris duplikat.",

            f"{'Mayoritas' if (sum(1 for s in num_stats if s.get('normality')=='Normal') > len(num_stats)/2) else 'Sebagian'} "
            "variabel numerik tidak mengikuti distribusi normal, sehingga analisis lanjutan "
            "perlu mempertimbangkan metode non-parametrik atau transformasi data.",

            f"Total {sum(int(s.get('outliers',0) or 0) for s in num_stats):,} pencilan terdeteksi "
            "yang memerlukan penanganan khusus sesuai konteks domain.",

            f"{len(insights)} insight otomatis berhasil diidentifikasi, "
            f"termasuk {sum(1 for i in insights if i.get('type') in ('warning','danger'))} "
            "temuan yang memerlukan tindakan segera.",
        ]

        for conc in conclusions:
            story.append(Paragraph(f"• {conc}", bullet_sty))

        story.append(Spacer(1, 0.3*cm))
        story.append(PageBreak())

        # ─────────────────────────────────────────────────────────────────
        # BAB IX — SARAN
        # ─────────────────────────────────────────────────────────────────
        story.append(Paragraph("Bab IX — Saran", h1_sty))
        story.append(HRFlowable(width="100%", thickness=1, color=C_TEAL, spaceAfter=10))

        story.append(Paragraph(
            "Berdasarkan hasil analisis eksplorasi data, berikut adalah saran-saran yang "
            "direkomendasikan untuk tahapan selanjutnya dalam alur kerja ilmu data:",
            body_sty
        ))

        suggestions = [
            ("Penanganan Nilai Hilang",
             f"{'Tidak diperlukan tindakan khusus karena tidak ada nilai hilang.' if total_missing == 0 else f'Terapkan strategi imputasi yang sesuai untuk {total_missing:,} nilai hilang. Gunakan SimpleImputer untuk data acak, KNNImputer untuk pola struktural, atau IterativeImputer (MICE) untuk dataset kompleks.'}"),

            ("Penanganan Duplikasi",
             f"{'Tidak ditemukan duplikasi.' if dup_rows == 0 else f'Hapus {dup_rows:,} baris duplikat menggunakan df.drop_duplicates() sebelum pemodelan untuk menghindari bias.'}"),

            ("Transformasi Data",
             "Untuk variabel dengan distribusi sangat miring (|skewness| > 2), "
             "terapkan transformasi logaritmik (np.log1p) atau Box-Cox untuk menormalkan distribusi "
             "sebelum digunakan dalam model berbasis asumsi normalitas."),

            ("Penanganan Pencilan",
             f"Analisis lebih lanjut diperlukan untuk {len(outlier_cols if 'outlier_cols' in dir() else [])} "
             "variabel yang mengandung pencilan. Gunakan pendekatan domain-driven untuk memutuskan "
             "apakah pencilan merupakan noise atau informasi berharga."),

            ("Feature Engineering",
             "Eksplorasi pembuatan fitur baru berdasarkan kombinasi variabel yang memiliki "
             "korelasi tinggi atau pola temporal (jika data time series tersedia)."),

            ("Pemilihan Algoritma",
             "Berdasarkan karakteristik data, pertimbangkan algoritma yang robust terhadap "
             "pencilan seperti Random Forest, Gradient Boosting, atau SVR daripada model "
             "linear standar."),
        ]

        for title_s, desc_s in suggestions:
            story.append(Paragraph(f"<b>{title_s}</b>", h3_sty))
            story.append(Paragraph(desc_s, body_sty))

        story.append(PageBreak())

        # ─────────────────────────────────────────────────────────────────
        # BAB X — LAMPIRAN (PRATINJAU DATA)
        # ─────────────────────────────────────────────────────────────────
        story.append(Paragraph("Bab X — Lampiran: Pratinjau Data (10 Baris Pertama)", h1_sty))
        story.append(HRFlowable(width="100%", thickness=1, color=C_TEAL, spaceAfter=10))
        story.append(Paragraph(
            "Lampiran berikut menampilkan 10 baris pertama dari dataset sebagai referensi "
            "struktur data mentah yang digunakan dalam analisis ini.",
            body_sty
        ))

        preview = self.df.head(10).copy()
        for col in preview.select_dtypes(include=["datetime","datetimetz"]).columns:
            preview[col] = preview[col].astype(str)
        preview = preview.fillna("—")

        # Batasi kolom agar muat di halaman A4
        max_cols_preview = 8
        if len(preview.columns) > max_cols_preview:
            story.append(Paragraph(
                f"Catatan: Hanya menampilkan {max_cols_preview} kolom pertama dari {len(preview.columns)} kolom total.",
                small_sty
            ))
            preview = preview.iloc[:, :max_cols_preview]

        pv_cols = list(preview.columns)
        cw_pv   = W / len(pv_cols)
        pv_rows = [pv_cols] + [[str(v)[:22] for v in row] for row in preview.values.tolist()]
        story.append(make_table(pv_rows, col_widths=[cw_pv]*len(pv_cols)))
        story.append(Paragraph("Tabel 6. Pratinjau 10 Baris Pertama Dataset", caption_sty))

        # Tipe data kolom
        story.append(Spacer(1, 0.4*cm))
        story.append(Paragraph("Tipe Data per Kolom", h2_sty))
        dtype_rows = [["Kolom","Tipe Data","Non-Null Count","Contoh Nilai"]]
        for col in list(self.df.columns)[:20]:
            sample = self.df[col].dropna().iloc[0] if self.df[col].notna().any() else "—"
            dtype_rows.append([
                col,
                str(self.df[col].dtype),
                f"{self.df[col].notna().sum():,}",
                str(sample)[:30]
            ])
        cw_dt = [4*cm, 2.5*cm, 2.5*cm, W-9*cm]
        story.append(make_table(dtype_rows, col_widths=cw_dt))
        story.append(Paragraph("Tabel 7. Tipe Data dan Informasi Kolom", caption_sty))

        if len(self.df.columns) > 20:
            story.append(Paragraph(
                f"Catatan: Hanya menampilkan 20 dari {len(self.df.columns)} kolom total.",
                small_sty
            ))

        # Footer
        story.append(Spacer(1, 0.8*cm))
        story.append(HRFlowable(width="100%", thickness=0.5, color=C_TEAL, spaceAfter=8))
        story.append(Paragraph(
            f"Auto EDA Insight  ·  SD-1306 Data Science Programming  ·  "
            f"Institut Teknologi Sains Bandung  ·  {self.now}",
            sty("FT", fontSize=7.5, textColor=C_GRAY, alignment=TA_CENTER)
        ))

        doc.build(story, canvasmaker=PageNumCanvas)
        return path
    # ══════════════════════════════════════════════════════════════════════════
    # EXCEL (multi-sheet, full dashboard)
    # ══════════════════════════════════════════════════════════════════════════
    def generate_excel(self, out_dir: str) -> str:
        path = os.path.join(out_dir, f"eda_report_{self.ts_str}.xlsx")
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            from openpyxl.styles import (PatternFill, Font, Alignment,
                                          Border, Side, numbers)
            from openpyxl.utils import get_column_letter

            wb = writer.book

            FILL_TEAL  = PatternFill("solid", fgColor="25C5E9")
            FILL_INK2  = PatternFill("solid", fgColor="061e35")
            FILL_INK3  = PatternFill("solid", fgColor="0a2e4e")
            FILL_GREEN = PatternFill("solid", fgColor="22c55e")
            FILL_RED   = PatternFill("solid", fgColor="ef4444")
            FILL_AMBER = PatternFill("solid", fgColor="f59e0b")

            FONT_HDR   = Font(bold=True, color="021225", name="Calibri", size=10)
            FONT_BODY  = Font(color="F2FFF6", name="Calibri", size=9)
            FONT_TEAL  = Font(bold=True, color="25C5E9", name="Calibri", size=9)
            FONT_TITLE = Font(bold=True, color="25C5E9", name="Calibri", size=14)

            ALIGN_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ALIGN_L = Alignment(horizontal="left",   vertical="center", wrap_text=True)

            thin = Side(style="thin", color="238689")
            BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

            def style_header_row(ws, row_num, n_cols, fill=FILL_TEAL):
                for col in range(1, n_cols+1):
                    cell = ws.cell(row=row_num, column=col)
                    cell.fill = fill
                    cell.font = FONT_HDR
                    cell.alignment = ALIGN_C
                    cell.border = BORDER

            def style_data_row(ws, row_num, n_cols, even=True):
                fill = FILL_INK2 if even else FILL_INK3
                for col in range(1, n_cols+1):
                    cell = ws.cell(row=row_num, column=col)
                    cell.fill = fill
                    cell.font = FONT_BODY
                    cell.alignment = ALIGN_L
                    cell.border = BORDER

            def auto_col_width(ws, min_w=8, max_w=40):
                for col_cells in ws.columns:
                    length = max(len(str(c.value or "")) for c in col_cells)
                    ws.column_dimensions[get_column_letter(col_cells[0].column)].width = \
                        min(max(length + 2, min_w), max_w)

            # ── Sheet 0: Cover ────────────────────────────────────────────────
            ws0 = wb.create_sheet("Cover")
            ws0.sheet_view.showGridLines = False
            ws0["A1"] = "Auto EDA Insight — Full Dashboard Report"
            ws0["A1"].font = FONT_TITLE
            ws0["A2"] = "SD-1306 Data Science Programming · Institut Teknologi Sains Bandung"
            ws0["A2"].font = Font(color="CAFFDE", name="Calibri", size=10)
            ws0["A4"] = "File"
            ws0["B4"] = self._filename()
            ws0["A5"] = "Rows"
            ws0["B5"] = len(self.df)
            ws0["A6"] = "Columns"
            ws0["B6"] = len(self.df.columns)
            ws0["A7"] = "Generated"
            ws0["B7"] = self.now
            for row in ws0.iter_rows(min_row=4, max_row=7):
                row[0].font = FONT_TEAL
                row[1].font = FONT_BODY
                for cell in row:
                    cell.fill = FILL_INK2
                    cell.border = BORDER

            # ── Sheet 1: Data Preview ─────────────────────────────────────────
            preview = self.df.head(100).copy()
            for col in preview.select_dtypes(include=["datetime","datetimetz"]).columns:
                preview[col] = preview[col].astype(str)
            preview = preview.fillna("")
            preview.to_excel(writer, sheet_name="Data Preview", index=False, startrow=1)
            ws1 = writer.sheets["Data Preview"]
            ws1["A1"] = f"Data Preview — {self._filename()} (first 100 rows)"
            ws1["A1"].font = Font(bold=True, color="25C5E9", name="Calibri", size=12)
            style_header_row(ws1, 2, len(preview.columns))
            for i in range(3, 3+len(preview)):
                style_data_row(ws1, i, len(preview.columns), even=(i%2==0))
            auto_col_width(ws1)

            # ── Sheet 2: Numerical Stats ──────────────────────────────────────
            num_stats = self._num_stats()
            if num_stats:
                num_keys   = ["column","count","mean","median","min","max","std","variance",
                               "mode","skewness","kurtosis","q1","q3",
                               "missing_count","missing_pct","normality","outliers"]
                num_labels = ["Column","Count","Mean","Median","Min","Max","Std Dev","Variance",
                               "Mode","Skewness","Kurtosis","Q1","Q3",
                               "Missing","Missing%","Normality","Outliers"]
                rows_n = [[s.get(k,"") for k in num_keys] for s in num_stats]
                df_n   = pd.DataFrame(rows_n, columns=num_labels)
                df_n.to_excel(writer, sheet_name="Numerical Stats", index=False, startrow=1)
                ws2 = writer.sheets["Numerical Stats"]
                ws2["A1"] = "Descriptive Statistics — Numerical Variables"
                ws2["A1"].font = Font(bold=True, color="25C5E9", name="Calibri", size=12)
                style_header_row(ws2, 2, len(num_labels))
                for i, s in enumerate(num_stats):
                    row_n = i + 3
                    style_data_row(ws2, row_n, len(num_labels), even=(i%2==0))
                    # color normality cell (col 16)
                    norm_cell = ws2.cell(row=row_n, column=16)
                    if s.get("normality") == "Normal":
                        norm_cell.fill = FILL_GREEN
                        norm_cell.font = Font(bold=True, color="021225", name="Calibri", size=9)
                    else:
                        norm_cell.fill = FILL_RED
                        norm_cell.font = Font(bold=True, color="F2FFF6", name="Calibri", size=9)
                    # color outliers (col 17)
                    out_cell = ws2.cell(row=row_n, column=17)
                    if s.get("outliers", 0) > 0:
                        out_cell.fill = FILL_AMBER
                        out_cell.font = Font(bold=True, color="021225", name="Calibri", size=9)
                auto_col_width(ws2)

            # ── Sheet 3: Categorical Stats ────────────────────────────────────
            cat_stats = self._cat_stats()
            if cat_stats:
                cat_keys   = ["column","count","unique","mode","mode_freq","mode_pct",
                               "missing_count","missing_pct"]
                cat_labels = ["Column","Count","Unique","Mode","Mode Freq","Mode %",
                               "Missing","Missing%"]
                rows_c = [[s.get(k,"") for k in cat_keys] for s in cat_stats]
                df_c   = pd.DataFrame(rows_c, columns=cat_labels)
                df_c.to_excel(writer, sheet_name="Categorical Stats", index=False, startrow=1)
                ws3 = writer.sheets["Categorical Stats"]
                ws3["A1"] = "Descriptive Statistics — Categorical Variables"
                ws3["A1"].font = Font(bold=True, color="25C5E9", name="Calibri", size=12)
                style_header_row(ws3, 2, len(cat_labels))
                for i in range(len(cat_stats)):
                    style_data_row(ws3, i+3, len(cat_labels), even=(i%2==0))
                auto_col_width(ws3)

            # ── Sheet 4: Auto Insights ────────────────────────────────────────
            insights = self._insights()
            if insights:
                ins_data = []
                for i, ins in enumerate(insights, 1):
                    text = ins.get("text","").replace("<b>","").replace("</b>","")
                    ins_data.append({
                        "#"     : i,
                        "Icon"  : insight_symbol(ins),
                        "Type"  : ins.get("type","info").upper(),
                        "Title" : ins.get("title",""),
                        "Detail": text,
                    })
                df_ins = pd.DataFrame(ins_data)
                df_ins.to_excel(writer, sheet_name="Auto Insights", index=False, startrow=1)
                ws4 = writer.sheets["Auto Insights"]
                ws4["A1"] = "Auto Insights — Intelligent Data Analysis"
                ws4["A1"].font = Font(bold=True, color="25C5E9", name="Calibri", size=12)
                style_header_row(ws4, 2, 5)
                for i, ins in enumerate(insights):
                    row_i = i + 3
                    style_data_row(ws4, row_i, 5, even=(i%2==0))
                    tp = ins.get("type","info")
                    type_cell = ws4.cell(row=row_i, column=3)
                    if tp == "success":
                        type_cell.fill = FILL_GREEN
                        type_cell.font = Font(bold=True, color="021225", name="Calibri", size=9)
                    elif tp in ("warning","danger"):
                        type_cell.fill = FILL_AMBER
                        type_cell.font = Font(bold=True, color="021225", name="Calibri", size=9)
                auto_col_width(ws4)

            # ── Sheet 5: Time Series ──────────────────────────────────────────
            ts = self._ts()
            if ts.get("has_timeseries"):
                dt_col = ts.get("date_column","")
                num_c  = self.df.select_dtypes(include="number").columns[:4].tolist()
                ts_rows = []
                if dt_col in self.df.columns:
                    df_s = self.df.sort_values(dt_col)
                    for c in num_c:
                        s = df_s[c].dropna()
                        if len(s) < 2: continue
                        first_half  = s.iloc[:len(s)//2].mean()
                        second_half = s.iloc[len(s)//2:].mean()
                        trend = "Upward ↑" if second_half > first_half else "Downward ↓"
                        ts_rows.append({
                            "Column"    : c,
                            "Date Col"  : dt_col,
                            "First Date": str(df_s[dt_col].dropna().iloc[0])[:10],
                            "Last Date" : str(df_s[dt_col].dropna().iloc[-1])[:10],
                            "Min"       : round(float(s.min()), 4),
                            "Max"       : round(float(s.max()), 4),
                            "Mean"      : round(float(s.mean()), 4),
                            "Std"       : round(float(s.std()), 4),
                            "Trend"     : trend,
                        })
                if ts_rows:
                    df_ts = pd.DataFrame(ts_rows)
                    df_ts.to_excel(writer, sheet_name="Time Series", index=False, startrow=1)
                    ws5 = writer.sheets["Time Series"]
                    ws5["A1"] = f"Time Series Analysis — Date column: {dt_col}"
                    ws5["A1"].font = Font(bold=True, color="25C5E9", name="Calibri", size=12)
                    style_header_row(ws5, 2, len(df_ts.columns))
                    for i in range(len(ts_rows)):
                        row_i = i + 3
                        style_data_row(ws5, row_i, len(df_ts.columns), even=(i%2==0))
                        trend_cell = ws5.cell(row=row_i, column=9)
                        if "Upward" in str(ts_rows[i].get("Trend","")):
                            trend_cell.fill = FILL_GREEN
                            trend_cell.font = Font(bold=True, color="021225", name="Calibri", size=9)
                        else:
                            trend_cell.fill = FILL_RED
                            trend_cell.font = Font(bold=True, color="F2FFF6", name="Calibri", size=9)
                    auto_col_width(ws5)

            # remove default "Sheet" if it exists
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

            # reorder: Cover first
            wb.move_sheet("Cover", offset=-len(wb.sheetnames)+1)

        return path

    # ══════════════════════════════════════════════════════════════════════════
    # CSV (full data + meta sheet via multiple CSV files zipped)
    # ══════════════════════════════════════════════════════════════════════════
    def generate_csv(self, out_dir: str) -> str:
        """Return a ZIP containing: data.csv, numerical_stats.csv,
        categorical_stats.csv, insights.csv, timeseries_summary.csv"""
        import zipfile

        zip_path = os.path.join(out_dir, f"eda_report_{self.ts_str}.zip")

        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            # 1. Full dataset
            df_copy = self.df.copy()
            for col in df_copy.select_dtypes(include=["datetime","datetimetz"]).columns:
                df_copy[col] = df_copy[col].astype(str)
            zf.writestr("01_dataset_full.csv", df_copy.to_csv(index=False))

            # 2. Numerical stats
            num = self._num_stats()
            if num:
                zf.writestr("02_numerical_stats.csv", pd.DataFrame(num).to_csv(index=False))

            # 3. Categorical stats
            cat = self._cat_stats()
            if cat:
                zf.writestr("03_categorical_stats.csv", pd.DataFrame(cat).to_csv(index=False))

            # 4. Insights
            ins = self._insights()
            if ins:
                ins_clean = [{"icon":i.get("icon",""), "type":i.get("type",""),
                               "title":i.get("title",""),
                               "text":i.get("text","").replace("<b>","").replace("</b>","")}
                             for i in ins]
                zf.writestr("04_auto_insights.csv", pd.DataFrame(ins_clean).to_csv(index=False))

            # 5. Time series summary
            ts = self._ts()
            if ts.get("has_timeseries"):
                dt_col = ts.get("date_column","")
                num_c  = self.df.select_dtypes(include="number").columns[:4].tolist()
                ts_rows = []
                if dt_col in self.df.columns:
                    df_s = self.df.sort_values(dt_col)
                    for c in num_c:
                        s = df_s[c].dropna()
                        if len(s) < 2: continue
                        ts_rows.append({
                            "column": c, "date_col": dt_col,
                            "first_date": str(df_s[dt_col].dropna().iloc[0])[:10],
                            "last_date" : str(df_s[dt_col].dropna().iloc[-1])[:10],
                            "min": round(float(s.min()),4), "max": round(float(s.max()),4),
                            "mean": round(float(s.mean()),4), "std": round(float(s.std()),4),
                            "trend": "Upward" if s.iloc[len(s)//2:].mean() > s.iloc[:len(s)//2].mean() else "Downward"
                        })
                if ts_rows:
                    zf.writestr("05_timeseries_summary.csv", pd.DataFrame(ts_rows).to_csv(index=False))

            # 6. README
            readme = (
                "Auto EDA Insight — Full Dashboard Export\n"
                f"Generated: {self.now}\n"
                f"File    : {self._filename()}\n"
                f"Rows    : {len(self.df):,}\n"
                f"Columns : {len(self.df.columns)}\n\n"
                "Files in this archive:\n"
                "  01_dataset_full.csv        — Complete dataset\n"
                "  02_numerical_stats.csv     — Descriptive stats (numerical)\n"
                "  03_categorical_stats.csv   — Descriptive stats (categorical)\n"
                "  04_auto_insights.csv       — Auto-generated insights\n"
                "  05_timeseries_summary.csv  — Time series summary (if available)\n"
            )
            zf.writestr("README.txt", readme)

        return zip_path

    # ══════════════════════════════════════════════════════════════════════════
    # HTML (rich full-dashboard report)
    # ══════════════════════════════════════════════════════════════════════════
    def generate_html(self, out_dir: str) -> str:
        num_stats = self._num_stats()
        cat_stats = self._cat_stats()
        insights  = self._insights()
        ts        = self._ts()

        def tbl(headers, rows_data, key_list):
            th = "".join(f"<th>{h}</th>" for h in headers)
            tbody = ""
            for r in rows_data:
                tds = "".join(f"<td>{html.escape(str(r.get(k,'—')))}</td>" for k in key_list)
                tbody += f"<tr>{tds}</tr>"
            return f"<table><thead><tr>{th}</tr></thead><tbody>{tbody or '<tr><td colspan=99>—</td></tr>'}</tbody></table>"

        # build insight cards
        ins_html = ""
        for ins in insights:
            text = ins.get("text","")
            tp   = ins.get("type","info")
            border = {"success":"#22c55e","warning":"#f59e0b","danger":"#ef4444"}.get(tp, TEAL)
            ins_html += (
                f'<tr>'
                f'<td style="text-align:center;color:{border}">{insight_svg(ins)}</td>'
                f'<td style="color:{border};font-weight:600">{html.escape(ins.get("title",""))}</td>'
                f'<td style="color:#CAFFDE">{text}</td>'
                f'</tr>'
            )

        # time series summary
        ts_html = ""
        if ts.get("has_timeseries"):
            dt_col = ts.get("date_column","")
            num_c  = self.df.select_dtypes(include="number").columns[:4].tolist()
            ts_rows_html = ""
            if dt_col in self.df.columns:
                df_s = self.df.sort_values(dt_col)
                for c in num_c:
                    s = df_s[c].dropna()
                    if len(s) < 2: continue
                    up = s.iloc[len(s)//2:].mean() > s.iloc[:len(s)//2].mean()
                    trend = f'<span style="color:{"#22c55e" if up else "#ef4444"}">{"↑ Upward" if up else "↓ Downward"}</span>'
                    fd = str(df_s[dt_col].dropna().iloc[0])[:10]
                    ld = str(df_s[dt_col].dropna().iloc[-1])[:10]
                    ts_rows_html += (
                        f"<tr><td>{html.escape(c)}</td>"
                        f"<td>{fd}</td><td>{ld}</td>"
                        f"<td>{s.min():.2f}</td><td>{s.max():.2f}</td><td>{s.mean():.2f}</td>"
                        f"<td>{trend}</td></tr>"
                    )
            ts_html = f"""
            <table><thead><tr>
              <th>Column</th><th>First Date</th><th>Last Date</th>
              <th>Min</th><th>Max</th><th>Mean</th><th>Trend</th>
            </tr></thead><tbody>{ts_rows_html or '<tr><td colspan=7>—</td></tr>'}</tbody></table>"""
        else:
            ts_html = "<p style='color:rgba(202,255,222,.5)'>Tidak ada kolom datetime yang terdeteksi.</p>"

        # preview table (first 10)
        preview_df = self.df.head(10).copy()
        for col in preview_df.select_dtypes(include=["datetime","datetimetz"]).columns:
            preview_df[col] = preview_df[col].astype(str)
        preview_df = preview_df.fillna("—")
        pv_th = "".join(f"<th>{html.escape(str(c))}</th>" for c in preview_df.columns)
        pv_tb = "".join(
            "<tr>" + "".join(f"<td>{html.escape(str(v))[:50]}</td>" for v in row) + "</tr>"
            for row in preview_df.values.tolist()
        )

        num_th = ["Column","Count","Mean","Median","Min","Max","Std","Variance","Mode",
                  "Skewness","Kurtosis","Q1","Q3","Missing","Missing%","Normality","Outliers"]
        num_k  = ["column","count","mean","median","min","max","std","variance","mode",
                  "skewness","kurtosis","q1","q3","missing_count","missing_pct","normality","outliers"]
        num_rows_html = ""
        for s in num_stats:
            cells = ""
            for k, h in zip(num_k, num_th):
                v = s.get(k,"—")
                if k == "normality":
                    color = "#22c55e" if v == "Normal" else "#ef4444"
                    cells += f'<td style="color:{color};font-weight:600">{v}</td>'
                elif k == "outliers" and int(v or 0) > 0:
                    cells += f'<td style="color:#f59e0b;font-weight:600">{v}</td>'
                else:
                    cells += f"<td>{html.escape(str(v))}</td>"
            num_rows_html += f"<tr>{cells}</tr>"

        cat_th = ["Column","Count","Unique","Mode","Mode Freq","Mode%","Missing","Missing%"]
        cat_k  = ["column","count","unique","mode","mode_freq","mode_pct","missing_count","missing_pct"]
        cat_rows_html = ""
        for s in cat_stats:
            cells = "".join(f"<td>{html.escape(str(s.get(k,'—')))}</td>" for k in cat_k)
            cat_rows_html += f"<tr>{cells}</tr>"

        body = f"""<!DOCTYPE html>
<html lang="id">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Auto EDA Insight — Full Dashboard Report</title>
<style>
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{font-family:'Segoe UI',sans-serif;background:#021225;color:#F2FFF6;padding:0}}
  .header{{background:linear-gradient(135deg,#061e35,#0a2e4e);border-bottom:2px solid #25C5E9;padding:28px 40px 24px}}
  .logo-row{{display:flex;align-items:center;gap:14px;margin-bottom:8px}}
  .logo-icon{{width:44px;height:44px;background:linear-gradient(135deg,#25C5E9,#238689);border-radius:12px;display:flex;align-items:center;justify-content:center}}
  .logo-icon svg{{width:24px;height:24px;stroke:white;fill:none;stroke-width:2}}
  .logo-title{{font-size:22px;font-weight:800;color:#25C5E9;letter-spacing:-0.5px}}
  .logo-sub{{font-size:12px;color:#CAFFDE;margin-top:2px}}
  .meta-row{{display:flex;gap:24px;flex-wrap:wrap;margin-top:16px}}
  .meta-chip{{background:rgba(37,197,233,0.1);border:1px solid rgba(37,197,233,0.25);border-radius:8px;padding:8px 16px}}
  .meta-chip .lbl{{font-size:10px;color:#CAFFDE;text-transform:uppercase;letter-spacing:1px}}
  .meta-chip .val{{font-size:16px;font-weight:700;color:#25C5E9}}
  .container{{max-width:1400px;margin:0 auto;padding:32px 40px}}
  .section{{background:rgba(37,197,233,0.04);border:1px solid rgba(37,197,233,0.15);border-radius:16px;padding:24px;margin-bottom:28px}}
  .section-head{{display:flex;align-items:center;gap:10px;margin-bottom:18px;padding-bottom:12px;border-bottom:1px solid rgba(37,197,233,0.15)}}
  .section-num{{width:28px;height:28px;background:#25C5E9;border-radius:8px;display:flex;align-items:center;justify-content:center;font-weight:800;font-size:13px;color:#021225;flex-shrink:0}}
  .section-head h2{{font-size:15px;font-weight:700;color:#25C5E9}}
  .overflow{{overflow-x:auto}}
  table{{width:100%;border-collapse:collapse;font-size:12px}}
  th{{background:rgba(37,197,233,0.2);color:#25C5E9;padding:10px 12px;text-align:left;font-weight:600;white-space:nowrap;border-bottom:2px solid rgba(37,197,233,0.3)}}
  td{{padding:8px 12px;border-bottom:1px solid rgba(37,197,233,0.07);color:#F2FFF6}}
  tr:nth-child(even) td{{background:rgba(37,197,233,0.03)}}
  tr:hover td{{background:rgba(37,197,233,0.07)}}
  .ins-table td{{vertical-align:top;padding:10px 14px}}
  .badge{{display:inline-block;padding:2px 8px;border-radius:100px;font-size:10px;font-weight:700}}
  .badge-normal{{background:rgba(34,197,94,.15);color:#22c55e}}
  .badge-notnormal{{background:rgba(239,68,68,.15);color:#ef4444}}
  .badge-warn{{background:rgba(245,158,11,.15);color:#f59e0b}}
  .footer{{text-align:center;padding:24px;color:rgba(202,255,222,.4);font-size:11px;border-top:1px solid rgba(37,197,233,.1);margin-top:8px}}
</style>
</head>
<body>
<div class="header">
  <div class="logo-row">
    <div class="logo-icon">
      <svg viewBox="0 0 24 24"><path d="M3 3v18h18"/><path d="M7 16l4-4 4 4 4-8"/></svg>
    </div>
    <div>
      <div class="logo-title">Auto EDA Insight</div>
      <div class="logo-sub">SD-1306 Data Science Programming · Institut Teknologi Sains Bandung</div>
    </div>
  </div>
  <div class="meta-row">
    <div class="meta-chip"><div class="lbl">File</div><div class="val">{html.escape(self._filename())}</div></div>
    <div class="meta-chip"><div class="lbl">Rows</div><div class="val">{len(self.df):,}</div></div>
    <div class="meta-chip"><div class="lbl">Columns</div><div class="val">{len(self.df.columns)}</div></div>
    <div class="meta-chip"><div class="lbl">Generated</div><div class="val">{self.now}</div></div>
  </div>
</div>

<div class="container">

  <div class="section">
    <div class="section-head"><div class="section-num">1</div><h2>Data Preview (first 10 rows)</h2></div>
    <div class="overflow">
      <table><thead><tr>{pv_th}</tr></thead><tbody>{pv_tb}</tbody></table>
    </div>
  </div>

  <div class="section">
    <div class="section-head"><div class="section-num">2</div><h2>Descriptive Statistics — Numerical Variables</h2></div>
    <div class="overflow">
      <table><thead><tr>{"".join(f"<th>{h}</th>" for h in num_th)}</tr></thead>
      <tbody>{num_rows_html or '<tr><td colspan=17 style="color:rgba(202,255,222,.5)">Tidak ada variabel numerik</td></tr>'}</tbody></table>
    </div>
  </div>

  <div class="section">
    <div class="section-head"><div class="section-num">3</div><h2>Descriptive Statistics — Categorical Variables</h2></div>
    <div class="overflow">
      <table><thead><tr>{"".join(f"<th>{h}</th>" for h in cat_th)}</tr></thead>
      <tbody>{cat_rows_html or '<tr><td colspan=8 style="color:rgba(202,255,222,.5)">Tidak ada variabel kategorik</td></tr>'}</tbody></table>
    </div>
  </div>

  <div class="section">
    <div class="section-head"><div class="section-num">4</div><h2>Auto Insights</h2></div>
    <div class="overflow">
      <table class="ins-table">
        <thead><tr><th style="width:50px">Icon</th><th style="width:240px">Title</th><th>Detail</th></tr></thead>
        <tbody>{ins_html or '<tr><td colspan=3 style="color:rgba(202,255,222,.5)">Belum ada insights</td></tr>'}</tbody>
      </table>
    </div>
  </div>

  <div class="section">
    <div class="section-head"><div class="section-num">5</div><h2>Time Series Analysis</h2></div>
    <div class="overflow">{ts_html}</div>
  </div>

</div>
<div class="footer">
  Auto EDA Insight · SD-1306 Data Science Programming · Institut Teknologi Sains Bandung · {self.now}
</div>
</body></html>"""

        path = os.path.join(out_dir, f"eda_dashboard_{self.ts_str}.html")
        with open(path, "w", encoding="utf-8") as f:
            f.write(body)
        return path
